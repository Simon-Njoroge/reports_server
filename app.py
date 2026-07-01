from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from dotenv import load_dotenv
import logging
import os
import io
from functools import partial
from datetime import datetime, timezone

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas as pdfcanvas


try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)

load_dotenv()

app = Flask(__name__)

CORS(app, origins='*', supports_credentials=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

HOSPITAL_BLUE_DARK = '#0A2F5A'
HOSPITAL_BLUE_PRIMARY = '#1A5276'
HOSPITAL_BLUE_MID = '#2E86C1'
HOSPITAL_BLUE_LIGHT = '#5DADE2'
HOSPITAL_BLUE_PALE = '#D4E6F1'
HOSPITAL_BLUE_WHITE = '#EBF5FB'

HOSPITAL_GREEN = '#27AE60'
HOSPITAL_RED = '#E74C3C'
HOSPITAL_GOLD = '#F39C12'
HOSPITAL_PURPLE = '#8E44AD'

COLOR_SUCCESS = '#27AE60'
COLOR_FAILED = '#E74C3C'
COLOR_PENDING = '#F39C12'

COLOR_TEXT = '#1A2F3E'
COLOR_MUTED = '#5D7B93'
COLOR_FAINT = '#8AAEC2'
COLOR_BORDER = '#B8D4E3'
COLOR_CARD_BG = '#F8FCFF'
COLOR_STRIPED = '#F0F8FF'

CARD_ACCENTS = [HOSPITAL_BLUE_PRIMARY, HOSPITAL_BLUE_MID, HOSPITAL_GREEN, HOSPITAL_BLUE_DARK]

PAGE_SIZE = A4
PAGE_WIDTH, PAGE_HEIGHT = PAGE_SIZE
HEADER_HEIGHT = 1.05 * inch
FOOTER_HEIGHT = 0.55 * inch
SIDE_MARGIN = 0.55 * inch

def get_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='StatementTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor(HOSPITAL_BLUE_PRIMARY),
        alignment=TA_CENTER,
        leading=24
    ))

    styles.add(ParagraphStyle(
        name='ReportSubtitle',
        parent=styles['Normal'],
        fontSize=11.5,
        textColor=colors.HexColor(COLOR_MUTED),
        alignment=TA_LEFT,
        leading=15
    ))

    styles.add(ParagraphStyle(
        name='MetaLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor(HOSPITAL_BLUE_PRIMARY),
        leading=11
    ))

    styles.add(ParagraphStyle(
        name='MetaValue',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor(COLOR_TEXT),
        leading=11
    ))

    styles.add(ParagraphStyle(
        name='SectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor(HOSPITAL_BLUE_PRIMARY),
        leading=18
    ))

    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=colors.HexColor(HOSPITAL_BLUE_DARK),
        spaceAfter=0.12 * inch,
        leading=16
    ))

    styles.add(ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        textColor=colors.white,
        alignment=TA_LEFT,
        leading=11
    ))

    styles.add(ParagraphStyle(
        name='TableCell',
        parent=styles['Normal'],
        fontSize=8.3,
        textColor=colors.HexColor(COLOR_TEXT),
        alignment=TA_LEFT,
        leading=10.5
    ))

    styles.add(ParagraphStyle(
        name='TableCellRight',
        parent=styles['TableCell'],
        alignment=TA_RIGHT,
        fontName='Helvetica-Bold',
    ))

    styles.add(ParagraphStyle(
        name='SummaryNumber',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor(HOSPITAL_BLUE_DARK),
        alignment=TA_CENTER,
        leading=19
    ))

    styles.add(ParagraphStyle(
        name='SummaryLabel',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor(COLOR_MUTED),
        alignment=TA_CENTER,
        leading=10
    ))

    styles.add(ParagraphStyle(
        name='ChipText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.6,
        textColor=colors.white,
        alignment=TA_CENTER,
        leading=8
    ))

    styles.add(ParagraphStyle(
        name='Disclaimer',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.HexColor(COLOR_FAINT),
        alignment=TA_CENTER,
        leading=10
    ))

    styles.add(ParagraphStyle(
        name='HospitalHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor(HOSPITAL_BLUE_MID),
        alignment=TA_LEFT,
        leading=12
    ))

    styles.add(ParagraphStyle(
        name='KPIValue',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor(HOSPITAL_BLUE_DARK),
        alignment=TA_CENTER,
        leading=24
    ))

    return styles

def safe_amount(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0

def format_currency(amount):
    return f"KES {safe_amount(amount):,.2f}"

def format_date(date_str):
    try:
        if date_str:
            if isinstance(date_str, datetime):
                return date_str.strftime('%d %b %Y')
            if isinstance(date_str, str):
                dt = datetime.fromisoformat(str(date_str).replace('Z', '+00:00'))
                return dt.strftime('%d %b %Y')
            return str(date_str)[:16]
        return ''
    except Exception:
        return str(date_str)[:16]

def status_color(status):
    status = (status or '').lower()
    if status in ['active', 'successful']:
        return colors.HexColor(COLOR_SUCCESS)
    elif status in ['failed', 'resigned', 'terminated']:
        return colors.HexColor(COLOR_FAILED)
    return colors.HexColor(COLOR_PENDING)

def make_status_chip(status, styles):
    label = (status or 'unknown').upper()
    chip = Table([[Paragraph(label, styles['ChipText'])]])
    chip.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), status_color(status)),
        ('ROUNDEDCORNERS', [9, 9, 9, 9]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    return chip

def make_summary_card(number_text, label_text, accent_hex, width, styles):
    card = Table(
        [[''], [Paragraph(number_text, styles['SummaryNumber'])],
         [Paragraph(label_text, styles['SummaryLabel'])]],
        colWidths=[width],
        rowHeights=[5, None, None]
    )
    card.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor(accent_hex)),
        ('BACKGROUND', (0, 1), (0, 2), colors.HexColor(COLOR_CARD_BG)),
        ('ROUNDEDCORNERS', [8, 8, 8, 8]),
        ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor(COLOR_BORDER)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (0, 1), 6),
        ('BOTTOMPADDING', (0, 1), (0, 1), 2),
        ('TOPPADDING', (0, 2), (0, 2), 0),
        ('BOTTOMPADDING', (0, 2), (0, 2), 8),
        ('TOPPADDING', (0, 0), (0, 0), 0),
        ('BOTTOMPADDING', (0, 0), (0, 0), 0),
    ]))
    return card

class StatementCanvas(pdfcanvas.Canvas):
    def __init__(self, *args, **kwargs):
        self.report_meta = kwargs.pop('report_meta', {})
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_header()
            self._draw_footer(total_pages)
            super().showPage()
        super().save()

    def _draw_header(self):
        self.saveState()
        
        self.setFillColor(colors.HexColor(HOSPITAL_BLUE_DARK))
        self.rect(0, PAGE_HEIGHT - HEADER_HEIGHT, PAGE_WIDTH, HEADER_HEIGHT, fill=1, stroke=0)
        
        self.setFillColor(colors.HexColor(HOSPITAL_BLUE_MID))
        self.rect(0, PAGE_HEIGHT - HEADER_HEIGHT - 3, PAGE_WIDTH, 3, fill=1, stroke=0)
        
        self.setFillColor(colors.white)
        self.setFont('Helvetica-Bold', 16)
        self.drawString(SIDE_MARGIN + 5, PAGE_HEIGHT - 0.45 * inch, '+')
        
        self.setFont('Helvetica-Bold', 18)
        self.drawString(SIDE_MARGIN + 30, PAGE_HEIGHT - 0.45 * inch, 'HEALTHCARE SYSTEM')
        
        self.setFont('Helvetica', 9)
        self.setFillColor(colors.HexColor('#D4E8F7'))
        self.drawString(SIDE_MARGIN + 30, PAGE_HEIGHT - 0.66 * inch,
                         self.report_meta.get('report_title', 'HR Report'))
        
        right_x = PAGE_WIDTH - SIDE_MARGIN
        self.setFont('Helvetica-Bold', 10)
        self.setFillColor(colors.white)
        self.drawRightString(right_x, PAGE_HEIGHT - 0.45 * inch,
                              self.report_meta.get('event_title', 'Human Resources'))
        self.setFont('Helvetica', 8)
        self.setFillColor(colors.HexColor('#D4E8F7'))
        self.drawRightString(right_x, PAGE_HEIGHT - 0.62 * inch,
                              f"Generated {self.report_meta.get('generated', '')}")
        
        self.restoreState()

    def _draw_footer(self, total_pages):
        self.saveState()
        
        self.setStrokeColor(colors.HexColor(HOSPITAL_BLUE_PRIMARY))
        self.setLineWidth(0.8)
        self.line(SIDE_MARGIN, FOOTER_HEIGHT - 0.18 * inch,
                   PAGE_WIDTH - SIDE_MARGIN, FOOTER_HEIGHT - 0.18 * inch)
        
        self.setFont('Helvetica', 7.5)
        self.setFillColor(colors.HexColor(COLOR_MUTED))
        self.drawString(SIDE_MARGIN, FOOTER_HEIGHT - 0.34 * inch,
                         'Confidential HR Report - Authorized Personnel Only')
        
        self.drawCentredString(PAGE_WIDTH / 2.0, FOOTER_HEIGHT - 0.34 * inch,
                                self.report_meta.get('event_title', 'Healthcare System'))
        
        self.drawRightString(PAGE_WIDTH - SIDE_MARGIN, FOOTER_HEIGHT - 0.34 * inch,
                              f'Page {self.getPageNumber()} of {total_pages}')
        
        self.restoreState()

def clean_phone(phone):
    if not phone:
        return ''
    phone = str(phone).strip()
    if phone.startswith('="') and phone.endswith('"'):
        phone = phone[2:-1]
    phone = phone.replace('"', '').replace("'", "")
    return phone

def load_employees_from_excel(file_path):
    employees = []
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        field_map = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'employee id' in header_lower or 'id' in header_lower:
                field_map['employee_id'] = idx
            elif 'full name' in header_lower or 'name' in header_lower:
                field_map['full_name'] = idx
            elif 'gender' in header_lower:
                field_map['gender'] = idx
            elif 'department' in header_lower:
                field_map['department'] = idx
            elif 'job title' in header_lower or 'title' in header_lower:
                field_map['job_title'] = idx
            elif 'hire date' in header_lower or 'date' in header_lower:
                field_map['hire_date'] = idx
            elif 'annual salary' in header_lower or 'salary' in header_lower:
                field_map['annual_salary'] = idx
            elif 'performance rating' in header_lower or 'rating' in header_lower:
                field_map['performance_rating'] = idx
            elif 'satisfaction score' in header_lower or 'satisfaction' in header_lower:
                field_map['satisfaction_score'] = idx
            elif 'employment status' in header_lower or 'status' in header_lower:
                field_map['employment_status'] = idx
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            employee = {
                'employee_id': str(row[field_map.get('employee_id', 0)] or '').strip(),
                'full_name': str(row[field_map.get('full_name', 1)] or '').strip(),
                'gender': str(row[field_map.get('gender', 2)] or '').strip(),
                'department': str(row[field_map.get('department', 3)] or '').strip(),
                'job_title': str(row[field_map.get('job_title', 4)] or '').strip(),
                'hire_date': row[field_map.get('hire_date', 5)],
                'annual_salary': safe_amount(row[field_map.get('annual_salary', 6)]),
                'performance_rating': str(row[field_map.get('performance_rating', 7)] or '').strip(),
                'satisfaction_score': str(row[field_map.get('satisfaction_score', 8)] or '').strip(),
                'employment_status': str(row[field_map.get('employment_status', 9)] or '').strip()
            }
            employees.append(employee)
        
        logger.info(f"Loaded {len(employees)} employees from {file_path}")
        return employees
    
    except Exception as e:
        logger.error(f"Error loading Excel: {str(e)}")
        raise

def load_payments_from_csv(file_path):
    payments = []
    
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as file:
            reader = csv.DictReader(file)
            for row in reader:
                payment = {
                    'id': row.get('ID', '').strip(),
                    'reference': row.get('Reference', '').strip(),
                    'customerName': row.get('Customer Name', '').strip(),
                    'customerEmail': row.get('Customer Email', '').strip(),
                    'customerPhone': clean_phone(row.get('Customer Phone', '')),
                    'eventTitle': row.get('Event Title', '').strip(),
                    'amount': safe_amount(row.get('Amount', 0)),
                    'currency': row.get('Currency', 'KES').strip(),
                    'status': row.get('Status', '').strip().lower(),
                    'provider': row.get('Provider', '').strip(),
                    'paidAt': row.get('Paid At', '').strip(),
                    'createdAt': row.get('Created At', '').strip()
                }
                payments.append(payment)
        
        logger.info(f"Loaded {len(payments)} payments from {file_path}")
        return payments
    
    except Exception as e:
        logger.error(f"Error loading CSV: {str(e)}")
        raise

def detect_and_load_file(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext in ['.csv']:
        return load_payments_from_csv(file_path)
    elif file_ext in ['.xlsx', '.xls']:
        return load_employees_from_excel(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_ext}. Please use .csv, .xlsx, or .xls")

def generate_hr_report(employees, event_title, report_title, include_summary=True):
    buffer = io.BytesIO()
    styles = get_styles()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=PAGE_SIZE,
        rightMargin=SIDE_MARGIN,
        leftMargin=SIDE_MARGIN,
        topMargin=HEADER_HEIGHT + 0.3 * inch,
        bottomMargin=FOOTER_HEIGHT + 0.15 * inch
    )

    generated_str = datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')
    report_meta = {
        'event_title': event_title,
        'report_title': report_title,
        'generated': generated_str,
    }

    story = []

    story.append(Paragraph("EMPLOYEE HR REPORT", styles['StatementTitle']))
    story.append(Spacer(1, 5))
    
    story.append(Paragraph("Human Resources Department", styles['HospitalHeader']))
    story.append(Spacer(1, 10))

    logo_path = "favicon.png"
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.2 * inch, height=1.2 * inch)
    else:
        logo = Paragraph("+", styles["MetaValue"])

    left_data = [
        [Paragraph("Generated By:", styles["MetaLabel"]), Paragraph("HR Management System", styles["MetaValue"])],
        [Paragraph("Department:", styles["MetaLabel"]), Paragraph("Human Resources", styles["MetaValue"])],
        [Paragraph("Generated On:", styles["MetaLabel"]), Paragraph(datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC"), styles["MetaValue"])],
        [Paragraph("Employees:", styles["MetaLabel"]), Paragraph(str(len(employees)), styles["MetaValue"])],
        [Paragraph("Report ID:", styles["MetaLabel"]), Paragraph(f"HR-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M')}", styles["MetaValue"])],
    ]

    left_table = Table(left_data, colWidths=[120, 180])
    left_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, -1), 10),
    ]))

    top = Table([[left_table, logo]], colWidths=[380, 100])
    top.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
    ]))

    story.append(top)
    story.append(Spacer(1, 20))

    if include_summary:
        story.append(Paragraph("EMPLOYEE SUMMARY", styles['SectionTitle']))
        story.append(Spacer(1, 8))

        total_employees = len(employees)
        active = sum(1 for e in employees if (e.get('employment_status') or '').lower() == 'active')
        resigned = sum(1 for e in employees if (e.get('employment_status') or '').lower() == 'resigned')
        terminated = sum(1 for e in employees if (e.get('employment_status') or '').lower() == 'terminated')
        total_salary = sum(safe_amount(e.get('annual_salary')) for e in employees)
        avg_salary = total_salary / total_employees if total_employees else 0
        
        departments = {}
        for e in employees:
            dept = e.get('department', 'Unknown')
            departments[dept] = departments.get(dept, 0) + 1
        
        dept_summary = ', '.join([f"{dept}: {count}" for dept, count in list(departments.items())[:5]])

        summary_rows = [
            ["Metric", "Value"],
            ["Total Employees", str(total_employees)],
            ["Active Employees", str(active)],
            ["Resigned", str(resigned)],
            ["Terminated", str(terminated)],
            ["Total Annual Salary", format_currency(total_salary)],
            ["Average Salary", format_currency(avg_salary)],
            ["Departments", dept_summary if dept_summary else "N/A"],
        ]

        summary_table = Table(summary_rows, colWidths=[300, 180])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(HOSPITAL_BLUE_WHITE)]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 25))

    story.append(Paragraph("EMPLOYEE DETAILS", styles['SectionTitle']))
    story.append(Spacer(1, 8))

    headers = ['Emp ID', 'Name', 'Department', 'Job Title', 'Status', 'Salary']
    table_data = [[Paragraph(h, styles['TableHeader']) for h in headers]]

    for emp in employees:
        status = emp.get('employment_status') or 'unknown'
        row = [
            Paragraph(str(emp.get('employee_id') or '')[:12], styles['TableCell']),
            Paragraph(str(emp.get('full_name') or '')[:22], styles['TableCell']),
            Paragraph(str(emp.get('department') or '')[:18], styles['TableCell']),
            Paragraph(str(emp.get('job_title') or '')[:20], styles['TableCell']),
            make_status_chip(status, styles),
            Paragraph(format_currency(emp.get('annual_salary')), styles['TableCellRight']),
        ]
        table_data.append(row)

    col_widths = [80, 130, 110, 120, 70, 90]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), "Helvetica-Bold"),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor(COLOR_BORDER)),
        ('LINEABOVE', (0, 1), (-1, 1), 1.5, colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor(COLOR_STRIPED)))

    table.setStyle(TableStyle(table_style))
    story.append(table)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "This report contains confidential employee information and should only be shared with authorized personnel.",
        styles['Disclaimer']
    ))

    doc.build(story, canvasmaker=partial(StatementCanvas, report_meta=report_meta))
    buffer.seek(0)
    return buffer

def generate_payment_report(payments, event_title, report_title, include_summary=True):
    buffer = io.BytesIO()
    styles = get_styles()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=PAGE_SIZE,
        rightMargin=SIDE_MARGIN,
        leftMargin=SIDE_MARGIN,
        topMargin=HEADER_HEIGHT + 0.3 * inch,
        bottomMargin=FOOTER_HEIGHT + 0.15 * inch
    )

    generated_str = datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')
    report_meta = {
        'event_title': event_title,
        'report_title': report_title,
        'generated': generated_str,
    }

    story = []

    story.append(Paragraph("PAYMENT STATEMENT", styles['StatementTitle']))
    story.append(Spacer(1, 5))
    
    story.append(Paragraph("Financial Services Department", styles['HospitalHeader']))
    story.append(Spacer(1, 10))

    logo_path = "favicon.png"
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.2 * inch, height=1.2 * inch)
    else:
        logo = Paragraph("+", styles["MetaValue"])

    left_data = [
        [Paragraph("Generated By:", styles["MetaLabel"]), Paragraph("Financial System", styles["MetaValue"])],
        [Paragraph("Department:", styles["MetaLabel"]), Paragraph("Finance", styles["MetaValue"])],
        [Paragraph("Generated On:", styles["MetaLabel"]), Paragraph(datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC"), styles["MetaValue"])],
        [Paragraph("Transactions:", styles["MetaLabel"]), Paragraph(str(len(payments)), styles["MetaValue"])],
        [Paragraph("Report ID:", styles["MetaLabel"]), Paragraph(f"FIN-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M')}", styles["MetaValue"])],
    ]

    left_table = Table(left_data, colWidths=[120, 180])
    left_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, -1), 10),
    ]))

    top = Table([[left_table, logo]], colWidths=[380, 100])
    top.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
    ]))

    story.append(top)
    story.append(Spacer(1, 20))

    if include_summary:
        story.append(Paragraph("FINANCIAL SUMMARY", styles['SectionTitle']))
        story.append(Spacer(1, 8))

        total_payments = len(payments)
        total_amount = sum(safe_amount(p.get('amount')) for p in payments)
        successful = sum(1 for p in payments if (p.get('status') or '').lower() == 'successful')
        failed = sum(1 for p in payments if (p.get('status') or '').lower() == 'failed')
        avg_amount = total_amount / total_payments if total_payments else 0

        summary_rows = [
            ["Metric", "Value"],
            ["Total Transactions", str(total_payments)],
            ["Successful Payments", str(successful)],
            ["Failed Payments", str(failed)],
            ["Total Revenue", format_currency(total_amount)],
            ["Average Payment", format_currency(avg_amount)],
        ]

        summary_table = Table(summary_rows, colWidths=[300, 180])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(HOSPITAL_BLUE_WHITE)]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 25))

    story.append(Paragraph("TRANSACTION DETAILS", styles['SectionTitle']))
    story.append(Spacer(1, 8))

    headers = ['Reference', 'Date', 'Customer', 'Status', 'Amount']
    table_data = [[Paragraph(h, styles['TableHeader']) for h in headers]]

    for payment in payments:
        status = payment.get('status') or 'unknown'
        row = [
            Paragraph(str(payment.get('reference') or '')[:20], styles['TableCell']),
            Paragraph(format_date(payment.get('paidAt')), styles['TableCell']),
            Paragraph(str(payment.get('customerName') or '')[:25], styles['TableCell']),
            make_status_chip(status, styles),
            Paragraph(format_currency(payment.get('amount')), styles['TableCellRight']),
        ]
        table_data.append(row)

    col_widths = [100, 100, 180, 70, 80]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), "Helvetica-Bold"),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor(COLOR_BORDER)),
        ('LINEABOVE', (0, 1), (-1, 1), 1.5, colors.HexColor(HOSPITAL_BLUE_PRIMARY)),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor(COLOR_STRIPED)))

    table.setStyle(TableStyle(table_style))
    story.append(table)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "This report contains confidential financial information and should only be shared with authorized personnel.",
        styles['Disclaimer']
    ))

    doc.build(story, canvasmaker=partial(StatementCanvas, report_meta=report_meta))
    buffer.seek(0)
    return buffer

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'service': 'report-generator',
        'version': '2.0.0',
        'timestamp': datetime.now(timezone.utc).isoformat()
    }), 200

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    try:
        data = request.json
        if not data or 'payments' not in data:
            return jsonify({'success': False, 'error': 'Payments data is required'}), 400
        payments = data.get('payments', [])
        if not payments:
            return jsonify({'success': False, 'error': 'No payments to generate report'}), 400

        pdf_buffer = generate_payment_report(
            payments=payments,
            event_title=data.get('eventTitle', 'Payment Report'),
            report_title=data.get('reportTitle', 'Payment Summary Report'),
            include_summary=data.get('includeSummary', True)
        )

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"payment_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generate-report-from-file', methods=['POST'])
def generate_report_from_file():
    try:
        data = request.json
        if not data or 'filePath' not in data:
            return jsonify({'success': False, 'error': 'filePath is required'}), 400
        
        file_path = data.get('filePath')
        event_title = data.get('eventTitle', 'Report')
        report_title = data.get('reportTitle', 'Summary Report')
        
        logger.info(f"Loading data from: {file_path}")
        records = detect_and_load_file(file_path)
        
        if not records:
            return jsonify({'success': False, 'error': 'No records found in file'}), 400
        
        logger.info(f"Loaded {len(records)} records, generating report...")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.xlsx', '.xls']:
            pdf_buffer = generate_hr_report(
                employees=records,
                event_title=event_title,
                report_title=report_title,
                include_summary=data.get('includeSummary', True)
            )
            filename = f"hr_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
        else:
            pdf_buffer = generate_payment_report(
                payments=records,
                event_title=event_title,
                report_title=report_title,
                include_summary=data.get('includeSummary', True)
            )
            filename = f"payment_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
        
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 404
    except ValueError as e:
        logger.error(f"Value error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Error generating report from file: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'

    print("=" * 60)
    print("Healthcare Report Server")
    print("=" * 60)
    print(f"URL: http://localhost:{port}")
    print(f"Health: http://localhost:{port}/health")
    print(f"Generate from JSON: POST http://localhost:{port}/api/generate-report")
    print(f"Generate from File: POST http://localhost:{port}/api/generate-report-from-file")
    print("=" * 60)

    app.run(host='0.0.0.0', port=port, debug=debug)